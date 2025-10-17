# BLOCCO 1

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os, csv, json, zipfile, shutil, re, unicodedata, subprocess
from fpdf import FPDF
import openpyxl
from xml.etree.ElementTree import Element, SubElement, ElementTree
from odf.opendocument import OpenDocumentSpreadsheet
from odf.table import Table, TableRow, TableCell
from odf.text import P
from openpyxl.worksheet.table import Table as XLTable, TableStyleInfo
from openpyxl.utils import get_column_letter

tutte_le_tabelle = {}

def pulisci_nome_file(nome):
    nome = unicodedata.normalize('NFKD', nome).encode('ASCII', 'ignore').decode()
    nome = re.sub(r'[^\w\s-]', '', nome)
    nome = nome.strip().replace(' ', '_')
    return nome

def carica_esempio():
    esempio = (
        "📦 Magazzino Centrale\n"
        "Prodotto\tQuantità\tPrezzo\n"
        "Matite\t100\t0.50\n"
        "Penne\t200\t1.20\n"
        "---\n"
        "🚛 Spedizioni in uscita\n"
        "Data\tDestinazione\tColli\n"
        "24/06/2025\tRoma\t3\n"
        "25/06/2025\tMilano\t5"
    )
    area_testo.delete("1.0", tk.END)
    area_testo.insert("1.0", esempio)

def chiedi_tipo_numeri(headers):
    risultato = {"scelta": None}

    def scegli(scelta):
        risultato["scelta"] = scelta
        finestra.destroy()

    finestra = tk.Toplevel()
    finestra.title("Come trattare i numeri?")
    finestra.geometry("400x220")
    finestra.resizable(False, False)
    finestra.grab_set()

    titolo = tk.Label(finestra, text="Come vuoi interpretare i numeri nei dati?", font=("Segoe UI", 10, "bold"))
    titolo.pack(pady=(20, 10))

    descrizione = tk.Label(finestra, text="Scegli una delle opzioni qui sotto:", font=("Segoe UI", 9))
    descrizione.pack(pady=(0, 15))

    frame_bottoni = ttk.Frame(finestra)
    frame_bottoni.pack(pady=5)

    ttk.Button(frame_bottoni, text="📎 Trattali tutti come TESTO", width=30, command=lambda: scegli("testo")).pack(pady=3)
    ttk.Button(frame_bottoni, text="🔢 Trattali tutti come NUMERI", width=30, command=lambda: scegli("numerico")).pack(pady=3)
    ttk.Button(frame_bottoni, text="⚖️ Scegli per COLONNA", width=30, command=lambda: scegli("misto")).pack(pady=3)

    finestra.wait_window()

    if risultato["scelta"] in {"testo", "numerico"}:
        return risultato["scelta"], set()

    selezioni = {}
    selettore = tk.Toplevel()
    selettore.title("Scegli colonne da trattare come testo")
    selettore.geometry("400x400")
    selettore.grab_set()

    tk.Label(selettore, text="Spunta le colonne da salvare come TESTO:", font=("Segoe UI", 10)).pack(pady=10)
    corpo = ttk.Frame(selettore)
    corpo.pack(expand=True, fill="both", padx=20)

    for h in headers:
        var = tk.BooleanVar(value=True)
        cb = ttk.Checkbutton(corpo, text=h, variable=var)
        cb.pack(anchor="w", pady=2)
        selezioni[h] = var

    conferma = tk.BooleanVar(value=False)

    def chiudi():
        conferma.set(True)
        selettore.destroy()

    ttk.Button(selettore, text="Conferma selezione", command=chiudi).pack(pady=10)
    selettore.wait_variable(conferma)

    colonne_testo = {h for h, var in selezioni.items() if var.get()}
    return "misto", colonne_testo

def interpreta_valore(col, val, strategia, colonne_testo):
    if strategia == "testo" or (strategia == "misto" and col in colonne_testo):
        return str(val)
    try:
        return float(val) if "." in str(val) else int(val)
    except:
        return str(val)
    
# BLOCCO 2

def genera_tabelle():
    for tab in notebook.tabs():
        notebook.forget(tab)
    global tutte_le_tabelle
    tutte_le_tabelle = {}

    area_testo.tag_remove("errore", "1.0", tk.END)

    testo = area_testo.get("1.0", tk.END).strip()
    if not testo:
        messagebox.showwarning("Nessun testo", "Inserisci del testo prima di generare le tabelle.")
        return

    blocchi = [b.strip() for b in testo.split("---") if b.strip()]
    if not blocchi:
        messagebox.showwarning("Formato non valido", "Nessuna sezione trovata. Usa '---' per separare le tabelle.")
        return

    errori_trovati = False

    for blocco in blocchi:
        righe = [r.strip() for r in blocco.split("\n") if r.strip()]
        if len(righe) < 3 or "\t" not in righe[1]:
            riga_problematica = righe[1] if len(righe) > 1 else righe[0]
            righe_testo = area_testo.get("1.0", tk.END).split("\n")
            for i, riga in enumerate(righe_testo, start=1):
                if riga.strip() == riga_problematica.strip():
                    start = f"{i}.0"
                    end = f"{i}.end"
                    area_testo.tag_add("errore", start, end)
                    area_testo.tag_config("errore", background="#ffcccc", foreground="#900000", underline=True)
                    area_testo.mark_set("insert", start)
                    area_testo.see("insert")
                    break

            messagebox.showerror(
                "Errore di Formattazione",
                f"❌ Errore nella riga:\n\n→ {riga_problematica}\n\n"
                "📌 Ogni sezione deve contenere almeno 3 righe:\n"
                "1️⃣ Titolo della tabella\n"
                "2️⃣ Intestazione con colonne separate da TAB (⇥)\n"
                "3️⃣ Una o più righe di dati\n\n"
                "💡 Esempio valido:\n"
                "Magazzino\n"
                "Prodotto⇥Quantità⇥Prezzo\n"
                "Matite⇥100⇥0.50"
            )
            errori_trovati = True
            continue

        titolo = righe[0].strip()
        intestazione = righe[1].split("\t")
        dati = [r.split("\t") for r in righe[2:]]
        tutte_le_tabelle[titolo] = [intestazione] + dati

        tab_frame = ttk.Frame(notebook)
        notebook.add(tab_frame, text=titolo)

        canvas = tk.Canvas(tab_frame)
        h_scroll = ttk.Scrollbar(tab_frame, orient="horizontal", command=canvas.xview)
        v_scroll = ttk.Scrollbar(tab_frame, orient="vertical", command=canvas.yview)
        canvas.configure(xscrollcommand=h_scroll.set, yscrollcommand=v_scroll.set)

        h_scroll.pack(side="bottom", fill="x")
        v_scroll.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        frame = ttk.Frame(canvas)
        canvas.create_window((0, 0), window=frame, anchor="nw")

        def aggiorna_scroll(event, c=canvas):
            c.configure(scrollregion=c.bbox("all"))
        frame.bind("<Configure>", aggiorna_scroll)

        for c, col in enumerate(intestazione):
            ttk.Label(frame, text=col, font=("Segoe UI", 10, "bold"),
                      background="#e0f0ff", relief="solid", padding=4).grid(row=0, column=c, sticky="nsew")
            frame.grid_columnconfigure(c, weight=1)

        for r, riga in enumerate(dati, 1):
            for c, val in enumerate(riga):
                ttk.Label(frame, text=val, relief="solid", padding=4).grid(row=r, column=c, sticky="nsew")

    if errori_trovati:
        messagebox.showwarning("Attenzione", "Alcune sezioni non sono state generate a causa di errori di formato.")

# BLOCCO 3

def esporta_dati():
    if not tutte_le_tabelle:
        messagebox.showwarning("Attenzione", "Genera prima le tabelle.")
        return

    formato = formato_var.get()
    estensione = f".{formato}"
    tipi_file = [(f"File {formato.upper()}", f"*.{formato}")]
    percorso = filedialog.asksaveasfilename(defaultextension=estensione, filetypes=tipi_file)
    if not percorso:
        return

    temp_dir = "temp_export"
    os.makedirs(temp_dir, exist_ok=True)

    strategia_numeri, colonne_testo = "testo", set()
    if formato in {"xlsx", "json", "xml", "ods"}:
        intestazioni_esempio = next(iter(tutte_le_tabelle.values()))[0]
        strategia_numeri, colonne_testo = chiedi_tipo_numeri(intestazioni_esempio)

    try:
        contenuti_generati = []

        if formato == "xlsx":
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = openpyxl.Workbook()
            for i, (titolo, righe) in enumerate(tutte_le_tabelle.items()):
                nome_foglio = pulisci_nome_file(titolo)[:31]
                ws = wb.create_sheet(title=nome_foglio) if i > 0 else wb.active
                ws.title = nome_foglio
                header, data = righe[0], righe[1:]

                # 🔷 Riga 1 – Titolo visivo
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(header))
                ws["A1"] = titolo
                ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
                ws["A1"].alignment = Alignment(horizontal="center")
                ws["A1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

                # 📋 Riga 2 – Intestazioni
                for col_idx, col_name in enumerate(header, start=1):
                    cell = ws.cell(row=2, column=col_idx, value=col_name)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

                # 🟦 Righe 3+ – Dati con bande alternate
                fill_even = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                fill_odd  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

                for row_idx, row in enumerate(data, start=3):
                    valori = [interpreta_valore(h, v, strategia_numeri, colonne_testo) for h, v in zip(header, row)]
                    for col_idx, val in enumerate(valori, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=val)
                        cell.fill = fill_even if row_idx % 2 == 0 else fill_odd
                        cell.alignment = Alignment(horizontal="center")

                # 📏 Larghezza automatica
                for col_idx, col_name in enumerate(header, start=1):
                    col_letter = get_column_letter(col_idx)
                    max_len = len(str(col_name))
                    for row in data:
                        if col_idx - 1 < len(row):
                            val = str(row[col_idx - 1])
                            max_len = max(max_len, len(val))
                    ws.column_dimensions[col_letter].width = max(12, min(max_len + 2, 50))

            wb.save(percorso)
            contenuti_generati = [os.path.basename(percorso)]

        else:
            for titolo, righe in tutte_le_tabelle.items():
                nome_file = pulisci_nome_file(titolo[:30])
                path = os.path.join(temp_dir, f"{nome_file}.{formato}")
                header, data = righe[0], righe[1:]

                if formato == "csv":
                    with open(path, "w", newline="", encoding="utf-8") as f:
                        writer = csv.writer(f, delimiter=";")
                        writer.writerow(header)
                        writer.writerows(data)

                elif formato == "json":
                    lista = []
                    for row in data:
                        item = {h: interpreta_valore(h, v, strategia_numeri, colonne_testo)
                                for h, v in zip(header, row)}
                        lista.append(item)
                    with open(path, "w", encoding="utf-8") as f:
                        json.dump(lista, f, ensure_ascii=False, indent=2)

                elif formato == "xml":
                    root = Element("Tabella", attrib={"titolo": titolo})
                    for row in data:
                        riga_el = SubElement(root, "Riga")
                        for h, v in zip(header, row):
                            campo = SubElement(riga_el, h)
                            campo.text = str(interpreta_valore(h, v, strategia_numeri, colonne_testo))
                    ElementTree(root).write(path, encoding="utf-8", xml_declaration=True)

                elif formato == "txt":
                    with open(path, "w", encoding="utf-8") as f:
                        f.write(titolo + "\n\n")
                        f.write("\t".join(header) + "\n")
                        for row in data:
                            f.write("\t".join(row) + "\n")

                elif formato == "html":
                    html = f"<h2>{titolo}</h2><table border='1'>"
                    html += "<tr>" + "".join(f"<th>{h}</th>" for h in header) + "</tr>"
                    for row in data:
                        html += "<tr>" + "".join(f"<td>{c}</td>" for c in row) + "</tr>"
                    html += "</table>"
                    with open(path, "w", encoding="utf-8") as f:
                        f.write(html)

                elif formato == "pdf":
                    class PDF(FPDF): pass
                    pdf = PDF(orientation="L", unit="mm", format="A4")
                    pdf.add_page()
                    pdf.set_font("Helvetica", "B", 14)
                    pdf.cell(0, 10, titolo, ln=True)
                    pdf.ln(4)
                    pdf.set_font("Helvetica", "B", 10)
                    for col in header:
                        pdf.cell(40, 8, col, border=1)
                    pdf.ln()
                    pdf.set_font("Helvetica", "", 9)
                    for row in data:
                        for cell in row:
                            pdf.cell(40, 8, cell, border=1)
                        pdf.ln()
                    pdf.output(path)

                elif formato == "ods":
                    doc = OpenDocumentSpreadsheet()
                    sheet = Table(name=nome_file)
                    for row in [header] + data:
                        tr = TableRow()
                        for h, v in zip(header, row):
                            valore = interpreta_valore(h, v, strategia_numeri, colonne_testo)
                            cell = TableCell()
                            cell.addElement(P(text=str(valore)))
                            tr.addElement(cell)
                        sheet.addElement(tr)
                    doc.spreadsheet.addElement(sheet)
                    doc.save(path)

                contenuti_generati.append(os.path.basename(path))

        # 📋 Anteprima finale
        anteprima = "\n".join(contenuti_generati)
        messagebox.showinfo("Anteprima dei file", f"Questi file sono stati generati:\n\n{anteprima}")

        # 📦 ZIP o copia in cartella
        if zip_var.get() and formato != "xlsx":
            zip_path = percorso if percorso.endswith(".zip") else os.path.splitext(percorso)[0] + ".zip"
            with zipfile.ZipFile(zip_path, "w") as zf:
                for f in contenuti_generati:
                    zf.write(os.path.join(temp_dir, f), f)
            messagebox.showinfo("Esportato!", f"Dati compressi ed esportati in:\n{zip_path}")
        elif formato != "xlsx":
            folder_name = os.path.splitext(os.path.basename(percorso))[0] + "_export"
            final_dir = os.path.join(os.path.dirname(percorso), folder_name)
            os.makedirs(final_dir, exist_ok=True)
            for f in contenuti_generati:
                shutil.copy(os.path.join(temp_dir, f), os.path.join(final_dir, f))
            messagebox.showinfo("Esportato!", f"Dati esportati in cartella:\n{final_dir}")

    except Exception as e:
        messagebox.showerror("Errore", f"Errore durante l'esportazione:\n{e}")
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)

# BLOCCO 4

def apri_cartella_finale():
    percorso = filedialog.askdirectory(title="Scegli la cartella da aprire")
    if percorso:
        try:
            subprocess.Popen(f'explorer "{percorso}"')
        except Exception as e:
            messagebox.showerror("Errore", f"Impossibile aprire la cartella:\n{e}")

def cancella_dati():
    testo_corrente = area_testo.get("1.0", tk.END).strip()
    tab_presenti = notebook.tabs()

    if not testo_corrente and not tab_presenti:
        messagebox.showwarning("Nessuna azione", "Non ci sono dati da cancellare.")
        return

    area_testo.delete("1.0", tk.END)
    for tab in tab_presenti:
        notebook.forget(tab)
    tutte_le_tabelle.clear()

def valida_input():
    area_testo.tag_remove("errore", "1.0", tk.END)
    area_testo.tag_remove("corretto", "1.0", tk.END)

    testo = area_testo.get("1.0", tk.END).strip()
    if not testo:
        messagebox.showwarning("Nessun testo", "Inserisci del testo prima di controllare il formato.")
        return

    blocchi = [b.strip() for b in testo.split("---") if b.strip()]
    if not blocchi:
        messagebox.showwarning("Formato non valido", "Nessuna sezione trovata. Usa '---' per separare le tabelle.")
        return

    errori = []
    righe_testo = area_testo.get("1.0", tk.END).split("\n")

    for i, blocco in enumerate(blocchi, 1):
        righe = [r.strip() for r in blocco.split("\n") if r.strip()]
        if len(righe) < 3:
            errori.append(f"❌ Sezione {i}: meno di 3 righe.")
            continue

        intestazione = righe[1]
        colonne_attese = intestazione.count("\t") + 1

        for riga_blocco in righe:
            riga_valida = False
            if riga_blocco == righe[0]:  # Titolo
                riga_valida = True
            elif riga_blocco == intestazione and "\t" in intestazione:
                riga_valida = True
            elif riga_blocco != intestazione and riga_blocco.count("\t") + 1 == colonne_attese:
                riga_valida = True
            else:
                errori.append(f"❌ Sezione {i}: riga non valida → '{riga_blocco}'")

            for j, riga_text in enumerate(righe_testo, start=1):
                if riga_text.strip() == riga_blocco.strip():
                    tag = "corretto" if riga_valida else "errore"
                    area_testo.tag_add(tag, f"{j}.0", f"{j}.end")
                    break

    area_testo.tag_config("corretto", background="#ccffcc", foreground="#006600")
    area_testo.tag_config("errore", background="#ffcccc", foreground="#900000", underline=True)

    if errori:
        messagebox.showwarning("Formato non valido", "\n\n".join(errori))
    else:
        messagebox.showinfo("Formato corretto", "Tutte le righe sono formattate correttamente.")

    # ⏳ Rimuove il verde dopo 3 secondi
    area_testo.after(3000, lambda: area_testo.tag_remove("corretto", "1.0", tk.END))

# ----- Finestra principale -----
finestra = tk.Tk()
finestra.title("Tabelle automatiche da testo")
finestra.geometry("1100x700")
finestra.configure(bg="#fdfdfd")

istruzioni = (
    "📌 COME USARE QUESTO PROGRAMMA:\n"
    "1. Incolla del testo con colonne separate da TAB (⇥).\n"
    "2. Ogni sezione deve iniziare con un titolo e usare --- come separatore.\n"
    "3. Premi 'Genera Tabelle' per visualizzarle in basso.\n"
    "4. Scegli il formato e premi 'Esporta'."
)

ttk.Label(finestra, text=istruzioni, justify="left", padding=10,
          background="#eef3ff", font=("Segoe UI", 10)).pack(fill="x", padx=10, pady=(10, 0))

area_testo = tk.Text(finestra, height=12, font=("Courier New", 10))
area_testo.pack(padx=10, pady=10, fill="x")

frame_bottoni = ttk.Frame(finestra)
frame_bottoni.pack(pady=5)

ttk.Button(frame_bottoni, text="📋 Carica esempio", command=carica_esempio, width=20).pack(side="left", padx=5)
ttk.Button(frame_bottoni, text="✅ Controlla Formato", command=valida_input, width=20).pack(side="left", padx=5)
ttk.Button(frame_bottoni, text="🔍 Genera Tabelle", command=genera_tabelle, width=20).pack(side="left", padx=5)
ttk.Button(frame_bottoni, text="💾 Esporta", command=esporta_dati, width=20).pack(side="left", padx=5)
ttk.Button(frame_bottoni, text="📂 Apri cartella", command=apri_cartella_finale, width=20).pack(side="left", padx=5)
ttk.Button(frame_bottoni, text="🧹 Cancella Dati", command=cancella_dati, width=20).pack(side="left", padx=5)

frame_formato = ttk.Frame(finestra)
frame_formato.pack(pady=5)

ttk.Label(frame_formato, text="Formato:").pack(side="left")

formato_var = tk.StringVar(value="xlsx")
formato_combobox = ttk.Combobox(
    frame_formato,
    textvariable=formato_var,
    values=["xlsx", "csv", "pdf", "json", "xml", "html", "txt", "ods"],
    state="readonly",
    width=8
)
formato_combobox.pack(side="left", padx=5)

zip_var = tk.BooleanVar(value=True)
chk_zip = ttk.Checkbutton(
    frame_formato,
    text="Comprimi in archivio ZIP",
    variable=zip_var
)
chk_zip.pack(side="left", padx=10)

notebook = ttk.Notebook(finestra)
notebook.pack(expand=1, fill="both", padx=10, pady=10)

finestra.mainloop()